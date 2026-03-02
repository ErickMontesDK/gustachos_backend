from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse

class ExcelGenerator:
    def __init__(self, sheet_name="Data"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name
        self.font_bold = Font(bold=True)
        self.alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color="2D6BE4", end_color="2D6BE4", fill_type="solid")
        self.font_bold_white = Font(bold=True, color="FFFFFF")
        self.fill_true = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        self.fill_false = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    def _get_value(self, obj, field):
        """Helper to get nested attributes or call methods."""
        if callable(field):
            return field(obj)
        
        parts = field.split('__')
        val = obj
        for part in parts:
            if val is None:
                return ""
            val = getattr(val, part, "")
        
        if isinstance(val, bool):
            return ("Yes", True) if val else ("No", False)
            
        return val

    def generate_excel(self, queryset, columns, filename="export.xlsx"):
        """
        columns: List of tuples (Header Name, Field/Callable)
        Example: [('ID', 'id'), ('Client', 'client__name'), ('Active', lambda x: 'Yes' if x.is_active else 'No')]
        """
        headers = [col[0] for col in columns]
        self.ws.append(headers)
        
        for cell in self.ws[1]:
            cell.fill = self.header_fill
            cell.font = self.font_bold_white
            cell.alignment = self.alignment
            cell.border = self.border

        for obj in queryset:
            row_values = [self._get_value(obj, col[1]) for col in columns]
            self.ws.append([v[0] if isinstance(v, tuple) else v for v in row_values])

            current_row = self.ws.max_row
            for idx, val in enumerate(row_values, start=1):
                if isinstance(val, tuple):
                    cell = self.ws.cell(row=current_row, column=idx)
                    cell.value = val[0]
                    cell.fill = self.fill_true if val[1] else self.fill_false

        for col in self.ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        
        self.wb.save(response)
        return response
